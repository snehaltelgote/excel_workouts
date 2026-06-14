import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd

# ---------------------------------------------------------
# STEP 1: Process, Match, and Calculate Differences
# ---------------------------------------------------------
def process_price_comparison(internal_file, competitor_file, output_file):
    print(f"Reading {internal_file} and {competitor_file}...")
    df_int = pd.read_excel(internal_file)
    df_comp = pd.read_excel(competitor_file)

    # Perform a left join to keep all internal catalog products
    print("Matching rows by product code and calculating metrics...")
    merged_df = pd.merge(df_int, df_comp, on='product_code', how='left')

    # Calculate Price Difference (Competitor - Internal)
    merged_df['price_difference'] = merged_df['competitor_price'] - merged_df['internal_price']

    # Calculate Margin Rate: (Competitor - Internal) / Competitor Price
    merged_df['margin_rate'] = merged_df['price_difference'] / merged_df['competitor_price']

    # Rename and reorder columns for a clean presentation
    merged_df = merged_df.rename(columns={
        'product_name': 'internal_name',
        'competitor_product_name': 'competitor_name'
    })

    columns_order = [
        'product_code', 'internal_name', 'competitor_name',
        'internal_price', 'competitor_price', 'price_difference', 'margin_rate'
    ]
    final_df = merged_df[columns_order]

    # Save base data to Excel
    final_df.to_excel(output_file, index=False)
    print(f"Base data exported to {output_file}.")


# ---------------------------------------------------------
# STEP 2: Formatting with OpenPyXL
# ---------------------------------------------------------
def format_excel_report(file_path):
    print("Applying professional formatting and conditional styling...")
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    ws.title = "Price Analysis"

    # Styles Definition
    font_family = "Segoe UI"
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid") # Dark Navy
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    data_font = Font(name=font_family, size=10)

    # Alert Fills for Margins
    green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid") # Soft Green
    red_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")   # Soft Red

    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    # Format Headers
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 28

    # Format Data Rows
    max_row = ws.max_row
    for row in range(2, max_row + 1):
        ws.row_dimensions[row].height = 20

        # Text alignments
        ws[f'A{row}'].alignment = Alignment(horizontal="center") # product_code
        ws[f'B{row}'].alignment = Alignment(horizontal="left")   # internal_name
        ws[f'C{row}'].alignment = Alignment(horizontal="left")   # competitor_name

        # Apply standard data font and borders
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
            cell = ws[f'{col}{row}']
            cell.font = data_font
            cell.border = thin_border

        # Numeric Formats (Currency)
        for col in ['D', 'E', 'F']:
            cell = ws[f'{col}{row}']
            cell.number_format = '"$"#,##0.00'
            cell.alignment = Alignment(horizontal="right")

        # Percentage Format (Margin)
        margin_cell = ws[f'G{row}']
        margin_cell.number_format = '0.00%'
        margin_cell.alignment = Alignment(horizontal="right")

        # Conditional Formatting based on price difference value
        diff_val = ws[f'F{row}'].value
        if diff_val is not None:
            if diff_val > 0:
                ws[f'F{row}'].fill = green_fill
                ws[f'G{row}'].fill = green_fill
            elif diff_val < 0:
                ws[f'F{row}'].fill = red_fill
                ws[f'G{row}'].fill = red_fill
        else:
            # Handle missing/unmatched rows cleanly
            ws[f'C{row}'].value = "Not Found"
            ws[f'C{row}'].font = Font(name=font_family, size=10, italic=True, color="7F7F7F")
            for col in ['E', 'F', 'G']:
                ws[f'{col}{row}'].value = "-"
                ws[f'{col}{row}'].alignment = Alignment(horizontal="center")

    # Auto-adjust column widths dynamically
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    wb.save(file_path)
    print("Formatting complete! Saved presentation-ready report.")


# ---------------------------------------------------------
# Execution Block
# ---------------------------------------------------------
if __name__ == "__main__":
    # Setup filenames
    file_a = "internal_catalog.xlsx"
    file_b = "competitor_prices.xlsx"
    file_out = "price_analysis_report.xlsx"

    # # Run pipeline
    process_price_comparison(file_a, file_b, file_out)
    format_excel_report(file_out)

    print(f"\nDemo process finished! Check your directory for: {file_out}")
